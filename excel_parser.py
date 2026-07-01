
#!/usr/bin/env python3
"""
Nessus Vulnerabilities HTML -> English GRC Excel Deliverables (v12)

Creates four English Excel files:
1. <PREFIX>_Asset_Inventory.xlsx
2. <PREFIX>_Threat_Catalogue.xlsx
3. <PREFIX>_Risk_Treatment_Plan.xlsx
4. <PREFIX>_Information_Security_Risk_Assessment.xlsx

Input expected:
- Tenable/Nessus plain "Vulnerabilities" HTML report (host sections with plugin summary tables)
- Optional Tenable/Nessus Operating Systems HTML report
- Optional Advanced IP Scanner HTML report
- Optional JSON risk overrides for organization-specific scoring

Example:
py -3.11 excel_parser.py --vulnerabilities "RAKIP_vulnerabilities.html" --os "RAKIP_operating_systems.html" --advip "RAKIP_adv-ip_scan.html" --out-prefix "RAKIP"
"""

from __future__ import annotations

import argparse
import re
import json
import time
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from pathlib import Path
from datetime import date, timedelta, datetime
from typing import Any, Optional

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -------------------------
# Severity and risk settings
# -------------------------
SEVERITY_COLORS = {
    "Critical": "8B0000",   # Dark red
    "High": "FF0000",       # Red
    "Medium": "FFA500",     # Orange
    "Low": "FFFF00",        # Yellow
    "Info": "9CC2E5",       # Light blue
    "None": "D9EAD3",       # Light green
    "Unknown": "D9D9D9",    # Grey
    "Informational": "9CC2E5",
}

SEVERITY_ORDER = {
    "Critical": 5,
    "High": 4,
    "Medium": 3,
    "Low": 2,
    "Info": 1,
    "None": 0,
    "Unknown": 0,
}

RISK_LEVEL_SCORES = {
    "Very Low": 1,
    "Low": 2,
    "Medium": 3,
    "High": 4,
    "Very High": 5,
}

PLUGIN_URL_PREFIX = "https://www.tenable.com/plugins/nessus/"
KNOWN_SEVERITIES = {"Critical", "High", "Medium", "Low", "Info", "None", "Unknown"}

# Optional organization-specific risk overrides are loaded only when the user
# supplies --risk-overrides. Without that option, the script applies the
# generic CVSS/severity-to-5x5 risk methodology below.



# -------------------------
# General utility functions
# -------------------------
def clean(value: Any, limit: int = 3000) -> str:
    """Normalize text and limit long Excel fields."""
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:limit]


def normalize_severity(value: Any) -> str:
    """Normalize Nessus severity text."""
    text = clean(value, 200).lower()
    if "critical" in text:
        return "Critical"
    if "high" in text:
        return "High"
    if "medium" in text:
        return "Medium"
    if "low" in text:
        return "Low"
    if "info" in text or "informational" in text:
        return "Info"
    if "none" in text:
        return "None"
    return "Unknown"


def severity_score(severity: Any) -> int:
    return SEVERITY_ORDER.get(normalize_severity(severity), 0)


def cvss_to_float(value: Any) -> Optional[float]:
    """Extract a CVSS numeric score from a Nessus field."""
    text = clean(value, 200).replace("*", "")
    match = re.search(r"\b(?:10(?:\.0)?|[0-9](?:\.[0-9])?)\b", text)
    if not match:
        return None
    try:
        score = float(match.group(0))
    except ValueError:
        return None
    if 0.0 <= score <= 10.0:
        return score
    return None


def likelihood_impact_from_cvss_or_severity(cvss: Any, severity: Any) -> tuple[str, int, str, int]:
    """
    Convert Nessus CVSS/severity into a 1-5 likelihood and impact matrix.

    Risk Score = Likelihood Score × Impact Score
    """
    score = cvss_to_float(cvss)
    sev = normalize_severity(severity)

    if score is not None:
        if score >= 9.0:
            return "Very High", 5, "Very High", 5
        if score >= 7.0:
            return "High", 4, "Very High", 5
        if score >= 4.0:
            return "Medium", 3, "Medium", 3
        if score > 0.0:
            return "Low", 2, "Low", 2
        return "Very Low", 1, "Very Low", 1

    fallback = {
        "Critical": ("Very High", 5, "Very High", 5),
        "High": ("High", 4, "Very High", 5),
        "Medium": ("Medium", 3, "Medium", 3),
        "Low": ("Low", 2, "Low", 2),
        "Info": ("Very Low", 1, "Very Low", 1),
        "None": ("Very Low", 1, "Very Low", 1),
    }
    return fallback.get(sev, ("Very Low", 1, "Very Low", 1))


def risk_rating_from_score(score: int) -> str:
    """Map a 1-25 score to the final risk rating."""
    if score >= 21:
        return "Critical"
    if score >= 13:
        return "High"
    if score >= 7:
        return "Medium"
    if score >= 4:
        return "Low"
    return "Informational"


def treatment_action_from_risk_rating(risk_rating: Any) -> str:
    rating = clean(risk_rating, 100)
    if rating == "Critical":
        return "Immediate remediation"
    if rating == "High":
        return "Prioritized remediation"
    if rating == "Medium":
        return "Planned remediation"
    if rating == "Low":
        return "Accept or remediate"
    return "Monitor"


def target_date_from_risk_rating(risk_rating: Any) -> str:
    """Suggested reassessment date based on calculated risk rating."""
    rating = clean(risk_rating, 100)
    days = {
        "Critical": 7,
        "High": 14,
        "Medium": 30,
        "Low": 60,
        "Informational": 90,
    }.get(rating, 90)
    return (date.today() + timedelta(days=days)).isoformat()


def read_soup(path: Path) -> BeautifulSoup:
    """Read HTML using lxml if installed, otherwise built-in html.parser."""
    html = path.read_text(encoding="utf-8", errors="ignore")
    try:
        return BeautifulSoup(html, "lxml")
    except Exception:
        return BeautifulSoup(html, "html.parser")


def load_risk_scoring_overrides(path: Optional[Path]) -> dict[str, tuple[str, int, str, int]]:
    """Load optional per-plugin likelihood/impact overrides from JSON.

    Expected JSON format:
    {
      "42873": {"likelihood": "Very High", "likelihood_score": 5,
                "impact": "Very Low", "impact_score": 1}
    }
    """
    if not path:
        return {}
    if not path.exists():
        raise SystemExit(f"[ERROR] Risk override file not found: {path}")

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"[ERROR] Could not read risk overrides: {exc}")

    if not isinstance(raw, dict):
        raise SystemExit("[ERROR] Risk overrides JSON must contain an object keyed by Nessus plugin ID.")

    valid_labels = set(RISK_LEVEL_SCORES)
    overrides: dict[str, tuple[str, int, str, int]] = {}
    for plugin_id, value in raw.items():
        if not str(plugin_id).isdigit() or not isinstance(value, dict):
            continue
        likelihood = clean(value.get("likelihood", ""), 100)
        impact = clean(value.get("impact", ""), 100)
        likelihood_score = value.get("likelihood_score")
        impact_score = value.get("impact_score")
        if (likelihood not in valid_labels or impact not in valid_labels or
                not isinstance(likelihood_score, int) or not isinstance(impact_score, int) or
                likelihood_score not in range(1, 6) or impact_score not in range(1, 6)):
            print(f"[!] Ignoring invalid risk override for plugin {plugin_id}.")
            continue
        overrides[str(plugin_id)] = (likelihood, likelihood_score, impact, impact_score)

    print(f"[+] Loaded {len(overrides)} organization-specific risk overrides.")
    return overrides


# -------------------------
# Parsers
# -------------------------


def extract_scan_date(path: Path) -> str:
    """Extract the Nessus report generation/scan date from the HTML header as YYYY-MM-DD."""
    soup = read_soup(path)
    text = soup.get_text(" ", strip=True)
    # Typical Nessus HTML: Fri, 12 Jun 2026 05:07:29 Pacific Standard Time
    match = re.search(r"(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun),\s+(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})", text)
    if match:
        try:
            return datetime.strptime(match.group(1), "%d %b %Y").date().isoformat()
        except ValueError:
            pass
    return ""


def extract_solution_from_plugin_html(html: str) -> str:
    """Extract one concise remediation sentence from a public Tenable plugin page."""
    soup = BeautifulSoup(html, "html.parser")
    heading = None
    for tag in soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6"]):
        if clean(tag.get_text(" ", strip=True), 100).lower() == "solution":
            heading = tag
            break
    if not heading:
        return ""

    parts = []
    for sibling in heading.find_next_siblings():
        if sibling.name in ["h1", "h2", "h3", "h4", "h5", "h6"]:
            break
        value = clean(sibling.get_text(" ", strip=True), 1000)
        if value:
            parts.append(value)
        if parts:
            break
    return clean(" ".join(parts), 1000)


def extract_description_from_plugin_html(html: str) -> str:
    """Extract the Tenable plugin Description text from a public plugin page."""
    soup = BeautifulSoup(html, "html.parser")
    heading = None
    for tag in soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6"]):
        if clean(tag.get_text(" ", strip=True), 100).lower() == "description":
            heading = tag
            break
    if not heading:
        return ""

    parts = []
    for sibling in heading.find_next_siblings():
        if sibling.name in ["h1", "h2", "h3", "h4", "h5", "h6"]:
            break
        value = clean(sibling.get_text(" ", strip=True), 3000)
        if value:
            parts.append(value)
    return clean(" ".join(parts), 3000)


def fetch_tenable_descriptions(plugin_ids: list[str], cache_path: Path = Path("tenable_description_cache.json")) -> dict[str, str]:
    """Fetch public Tenable plugin description text once per unique plugin ID and cache it locally."""
    cache: dict[str, str] = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            cache = {}

    unique_ids = sorted({str(pid) for pid in plugin_ids if str(pid).isdigit()})
    missing = [pid for pid in unique_ids if not cache.get(pid)]
    fallback = "Review the Tenable plugin reference for the detailed vulnerability description."
    if missing:
        print(f"[+] Retrieving Tenable description text for {len(missing)} unique plugins...")

        try:
            probe = Request(f"{PLUGIN_URL_PREFIX}{missing[0]}", headers={"User-Agent": "Mozilla/5.0 (compatible; Nessus-Lab-Report-Generator/1.0)"})
            with urlopen(probe, timeout=8) as response:
                first_html = response.read().decode("utf-8", errors="ignore")
            cache[missing[0]] = extract_description_from_plugin_html(first_html) or fallback
            missing = missing[1:]
        except (HTTPError, URLError, TimeoutError, OSError):
            print("[!] Tenable plugin pages are not reachable. Using fallback description text for this run.")
            for plugin_id in missing:
                cache[plugin_id] = fallback
            missing = []

    for index, plugin_id in enumerate(missing, start=1):
        url = f"{PLUGIN_URL_PREFIX}{plugin_id}"
        try:
            request = Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; Nessus-Lab-Report-Generator/1.0)"})
            with urlopen(request, timeout=12) as response:
                html = response.read().decode("utf-8", errors="ignore")
            description = extract_description_from_plugin_html(html)
            cache[plugin_id] = description or fallback
        except (HTTPError, URLError, TimeoutError, OSError):
            cache[plugin_id] = fallback
        if index < len(missing):
            time.sleep(0.2)

    try:
        cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return {pid: cache.get(pid, fallback) for pid in unique_ids}


def fetch_tenable_solutions(plugin_ids: list[str], cache_path: Path = Path("tenable_solution_cache.json")) -> dict[str, str]:
    """Fetch public Tenable plugin remediation text once per unique plugin ID and cache it locally."""
    cache: dict[str, str] = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            cache = {}

    unique_ids = sorted({str(pid) for pid in plugin_ids if str(pid).isdigit()})
    missing = [pid for pid in unique_ids if not cache.get(pid)]
    fallback = "Review the Tenable plugin reference and apply the vendor-recommended remediation."
    if missing:
        print(f"[+] Retrieving Tenable remediation text for {len(missing)} unique plugins...")

        # Test connectivity once. If Tenable cannot be reached, do not wait on every plugin page.
        try:
            probe = Request(f"{PLUGIN_URL_PREFIX}{missing[0]}", headers={"User-Agent": "Mozilla/5.0 (compatible; Nessus-Lab-Report-Generator/1.0)"})
            with urlopen(probe, timeout=8) as response:
                first_html = response.read().decode("utf-8", errors="ignore")
            cache[missing[0]] = extract_solution_from_plugin_html(first_html) or fallback
            missing = missing[1:]
        except (HTTPError, URLError, TimeoutError, OSError):
            print("[!] Tenable plugin pages are not reachable. Using fallback remediation text for this run.")
            for plugin_id in missing:
                cache[plugin_id] = fallback
            missing = []

    for index, plugin_id in enumerate(missing, start=1):
        url = f"{PLUGIN_URL_PREFIX}{plugin_id}"
        try:
            request = Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; Nessus-Lab-Report-Generator/1.0)"})
            with urlopen(request, timeout=12) as response:
                html = response.read().decode("utf-8", errors="ignore")
            solution = extract_solution_from_plugin_html(html)
            cache[plugin_id] = solution or fallback
        except (HTTPError, URLError, TimeoutError, OSError):
            cache[plugin_id] = fallback
        if index < len(missing):
            time.sleep(0.2)

    try:
        cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return {pid: cache.get(pid, "Review the Tenable plugin reference and apply the vendor-recommended remediation.") for pid in unique_ids}


def parse_nessus_vulnerabilities(path: Path) -> pd.DataFrame:
    """
    Parse the plain Tenable/Nessus Vulnerabilities HTML report.

    This report is usually grouped by host and includes plugin summary rows with:
    Severity, CVSS, VPR Score, EPSS Score, Plugin ID, Plugin Name.

    It normally does not include full Description/Solution text, so the treatment plan
    references the Tenable plugin URL for remediation details.
    """
    soup = read_soup(path)

    # Nessus uses a 22px <div> as the host heading. Depending on the scan,
    # that heading can be an IP address, a hostname, or a FQDN. Do not require
    # an IPv4-only pattern here, otherwise hostname-based sections are skipped.
    host_divs = []
    for div in soup.find_all("div"):
        text = clean(div.get_text(" ", strip=True), 300)
        style = div.get("style", "") or ""
        if text and "font-size: 22px" in style:
            host_divs.append(div)

    rows = []

    for host_div in host_divs:
        host = clean(host_div.get_text(" ", strip=True), 100)

        for sibling in host_div.find_next_siblings():
            if sibling.name == "h6":
                break

            sibling_style = sibling.get("style", "") or ""
            if sibling.name == "div" and "font-size: 22px" in sibling_style:
                break

            if sibling.name != "div" or "table-wrapper" not in (sibling.get("class") or []):
                continue

            # The hidden detailed table contains rows with class plugin-row.
            for tr in sibling.find_all("tr", class_="plugin-row"):
                values = [clean(td.get_text(" ", strip=True), 1000) for td in tr.find_all("td")]
                values = [v for v in values if v]

                # Expected meaningful values:
                # [Severity, CVSS v3.0, VPR Score, EPSS Score, Plugin ID, Plugin Name]
                if len(values) < 6:
                    continue

                severity = normalize_severity(values[0])
                if severity not in KNOWN_SEVERITIES or severity == "Unknown":
                    continue

                cvss = values[1]
                vpr = values[2]
                epss = values[3]
                plugin_id = values[4]
                plugin_name = values[5]

                if not re.fullmatch(r"\d+", plugin_id):
                    continue

                plugin_url = f"{PLUGIN_URL_PREFIX}{plugin_id}"

                rows.append({
                    "Source File": path.name,
                    "Host": host,
                    "Operating System": "",
                    "Plugin ID": plugin_id,
                    "Plugin URL": plugin_url,
                    "Name": plugin_name,
                    "Severity": severity,
                    "CVSS": cvss,
                    "VPR Score": vpr,
                    "EPSS Score": epss,
                    "CVE": "",
                    "Port / Protocol": "",
                    "Synopsis": plugin_name,
                    "Description": "Summary-only Nessus Vulnerabilities report. Review the Tenable plugin reference for detailed description and affected service information.",
                    "Solution": "Review the Tenable plugin reference and apply the vendor-recommended remediation or configuration change.",
                    "Plugin Output": "",
                })

    columns = [
        "Source File", "Host", "Operating System", "Plugin ID", "Plugin URL", "Name", "Severity", "CVSS",
        "VPR Score", "EPSS Score", "CVE", "Port / Protocol", "Synopsis", "Description", "Solution", "Plugin Output"
    ]

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=columns)

    # Keep every row reported by Nessus. The same plugin can legitimately appear
    # more than once for a host when it affects different services/instances.
    # Do not deduplicate at this stage.
    df["_severity_score"] = df["Severity"].map(severity_score)
    df["_cvss_score"] = df["CVSS"].map(lambda x: cvss_to_float(x) if cvss_to_float(x) is not None else -1)
    df = df.sort_values(["_severity_score", "_cvss_score", "Host", "Plugin ID"], ascending=[False, False, True, True])
    df = df.drop(columns=["_severity_score", "_cvss_score"])
    return df[columns]


def parse_operating_systems(path: Optional[Path]) -> pd.DataFrame:
    """Parse optional Nessus Operating Systems HTML report."""
    if not path or not path.exists():
        return pd.DataFrame()

    soup = read_soup(path)
    rows = []

    for tr in soup.find_all("tr", class_="plugin-row"):
        cells = [clean(td.get_text(" ", strip=True), 1000) for td in tr.find_all("td")]
        cells = [cell for cell in cells if cell]
        if len(cells) >= 4 and any(x in cells[0].lower() for x in ["windows", "linux", "nsx", "trellix", "unknown"]):
            rows.append({
                "Operating System": cells[0],
                "Count": cells[1],
                "Unsupported": cells[2],
                "Hosts": cells[3],
            })

    return pd.DataFrame(rows).drop_duplicates()


def build_host_os_map(os_df: pd.DataFrame) -> dict[str, str]:
    """Create a Host -> Operating System lookup from the OS report."""
    host_os: dict[str, str] = {}
    if os_df.empty:
        return host_os

    for _, row in os_df.iterrows():
        os_name = clean(row.get("Operating System", ""), 1000)
        hosts = clean(row.get("Hosts", ""), 5000)
        for host in re.findall(r"(?:\d{1,3}\.){3}\d{1,3}", hosts):
            host_os[host] = os_name
    return host_os


def enrich_findings_with_os(findings: pd.DataFrame, os_df: pd.DataFrame) -> pd.DataFrame:
    if findings.empty or os_df.empty:
        return findings
    host_os = build_host_os_map(os_df)
    findings = findings.copy()
    findings["Operating System"] = findings.apply(
        lambda row: row.get("Operating System", "") or host_os.get(str(row.get("Host", "")), ""),
        axis=1,
    )
    return findings


def parse_advanced_ip_scanner(path: Optional[Path]) -> pd.DataFrame:
    """Parse optional Advanced IP Scanner HTML report."""
    if not path or not path.exists():
        return pd.DataFrame()

    soup = read_soup(path)
    rows = []

    for tr in soup.find_all("tr", class_=re.compile(r"r[12]")):
        cells = [clean(td.get_text(" ", strip=True), 500) for td in tr.find_all("td", class_="head")]
        if len(cells) >= 6 and cells[0].lower() == "on":
            rows.append({
                "Status": cells[0],
                "Name": cells[1],
                "IP": cells[2],
                "Manufacturer": cells[3],
                "MAC Address": cells[4],
                "Comments": cells[5],
            })

    return pd.DataFrame(rows).drop_duplicates()


# -------------------------
# Workbook builders
# -------------------------
def build_risk_assessment(
    findings: pd.DataFrame,
    description_map: dict[str, str],
    risk_overrides: Optional[dict[str, tuple[str, int, str, int]]] = None,
) -> pd.DataFrame:
    """Build the streamlined English Information Security Risk Assessment sheet."""
    actionable = findings[findings["Severity"].isin(["Critical", "High", "Medium", "Low"])].copy()
    rows = []

    for _, finding in actionable.iterrows():
        severity = finding["Severity"]
        risk_id = len(rows) + 1
        plugin_id = str(finding.get("Plugin ID", ""))
        risk_overrides = risk_overrides or {}
        likelihood, likelihood_score, impact, impact_score = risk_overrides.get(
            plugin_id,
            likelihood_impact_from_cvss_or_severity(finding.get("CVSS", ""), severity),
        )
        risk_score = likelihood_score * impact_score
        risk_rating = risk_rating_from_score(risk_score)

        rows.append({
            "No.": risk_id,
            "Asset / Host": finding["Host"],
            "Plugin ID": finding["Plugin ID"],
            "Plugin URL": finding.get("Plugin URL", ""),
            "Vulnerability": finding.get("Synopsis", "") or finding["Name"],
            "Description": description_map.get(str(finding["Plugin ID"]), finding.get("Description", "")),
            "CVSS": finding.get("CVSS", ""),
            "Impact Score": impact_score,
            "Risk Score (Likelihood x Impact)": risk_score,
            "Risk Rating": risk_rating,
        })

    return pd.DataFrame(rows)


def build_risk_treatment_plan(
    risk_df: pd.DataFrame,
    findings: pd.DataFrame,
    solution_map: dict[str, str],
    scan_date: str,
) -> pd.DataFrame:
    """Build the Risk Treatment Plan with solution, scan date, and blank reassessment fields."""
    rows = []

    for _, risk in risk_df.iterrows():
        plugin_id = str(risk["Plugin ID"])
        rows.append({
            "Related Risk Assessment No.": risk["No."],
            "Asset / Host": risk["Asset / Host"],
            "Vulnerability Name": risk["Vulnerability"],
            "Plugin ID": plugin_id,
            "Plugin URL": risk.get("Plugin URL", ""),
            "Risk Rating": risk["Risk Rating"],
            "Severity": risk["Risk Rating"],
            "Risk Score (Likelihood x Impact)": risk["Risk Score (Likelihood x Impact)"],
            "Solution": solution_map.get(plugin_id, "Review the Tenable plugin reference and apply the vendor-recommended remediation."),
            "Date": scan_date,
            "Reassessment": "",
            "Reassessment Date": "",
        })

    return pd.DataFrame(rows)


def build_risk_criteria() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "Nessus Severity / CVSS Range": "Critical / 9.0-10.0",
            "Likelihood": "Very High",
            "Impact": "Very High",
            "Impact Score": 5,
            "Risk Calculation": "5 x 5",
            "Risk Score (Likelihood x Impact)": 25,
            "Risk Rating": "Critical",
            "Treatment Priority": "Immediate remediation",
            "Suggested Reassessment Window": "7 days",
        },
        {
            "Nessus Severity / CVSS Range": "High / 7.0-8.9",
            "Likelihood": "High",
            "Impact": "Very High",
            "Impact Score": 5,
            "Risk Calculation": "4 x 5",
            "Risk Score (Likelihood x Impact)": 20,
            "Risk Rating": "High",
            "Treatment Priority": "Prioritized remediation",
            "Suggested Reassessment Window": "14 days",
        },
        {
            "Nessus Severity / CVSS Range": "Medium / 4.0-6.9",
            "Likelihood": "Medium",
            "Impact": "Medium",
            "Impact Score": 3,
            "Risk Calculation": "3 x 3",
            "Risk Score (Likelihood x Impact)": 9,
            "Risk Rating": "Medium",
            "Treatment Priority": "Planned remediation",
            "Suggested Reassessment Window": "30 days",
        },
        {
            "Nessus Severity / CVSS Range": "Low / 0.1-3.9",
            "Likelihood": "Low",
            "Impact": "Low",
            "Impact Score": 2,
            "Risk Calculation": "2 x 2",
            "Risk Score (Likelihood x Impact)": 4,
            "Risk Rating": "Low",
            "Treatment Priority": "Accept or remediate",
            "Suggested Reassessment Window": "60 days",
        },
        {
            "Nessus Severity / CVSS Range": "Info / 0.0",
            "Likelihood": "Very Low",
            "Impact": "Very Low",
            "Impact Score": 1,
            "Risk Calculation": "1 x 1",
            "Risk Score (Likelihood x Impact)": 1,
            "Risk Rating": "Informational",
            "Treatment Priority": "Monitor",
            "Suggested Reassessment Window": "90 days",
        },
    ])


def build_threat_catalogue(findings: pd.DataFrame) -> pd.DataFrame:
    """Build a concise threat catalogue from the plain Nessus Vulnerabilities report."""
    if findings.empty:
        return pd.DataFrame(columns=["ID", "IP", "Severity", "Plugin Link", "Plugin Name"])

    rows = []
    catalogue = findings[findings["Severity"].isin(["Critical", "High", "Medium", "Low"])].copy()
    catalogue["_severity_score"] = catalogue["Severity"].map(severity_score)
    catalogue["_cvss_score"] = catalogue["CVSS"].map(lambda x: cvss_to_float(x) if cvss_to_float(x) is not None else -1)
    catalogue = catalogue.sort_values(["_severity_score", "_cvss_score", "Host", "Plugin ID"], ascending=[False, False, True, True])

    for _, finding in catalogue.iterrows():
        rows.append({
            "ID": len(rows) + 1,
            "IP": clean(finding.get("Host", ""), 200),
            "Severity": normalize_severity(finding.get("Severity", "")),
            "Plugin Link": finding.get("Plugin URL", "") or f"{PLUGIN_URL_PREFIX}{finding.get('Plugin ID', '')}",
            "Plugin Name": finding.get("Name", ""),
        })

    return pd.DataFrame(rows, columns=["ID", "IP", "Severity", "Plugin Link", "Plugin Name"])



def classify_asset(name: Any, ip: Any, operating_system: Any, manufacturer: Any) -> tuple[str, str]:
    """Return (Asset Classification, Threat Category) for the Asset Inventory."""
    n = clean(name, 300).lower()
    os_name = clean(operating_system, 500).lower()
    m = clean(manufacturer, 300).lower()

    if any(x in m for x in ["sophos", "fujitsu", "a10", "hewlett packard", "cisco", "tp-link", "shenzhen"]):
        return "Network Device", "Network Device Risk"
    if any(x in os_name for x in ["router", "switch", "firewall", "a10", "advanced core", "irmc"]):
        return "Network Device", "Network Device Risk"
    if any(x in n for x in ["dc", "server", "backup", "vserver", "dbserver"]):
        return "Server", "Server Risk"
    if any(x in os_name for x in ["windows server", "linux"]):
        return "Server", "Server Risk"
    if any(x in os_name for x in ["windows", "desktop"]):
        return "Endpoint", "Endpoint Device Risk"
    if any(x in os_name for x in ["iphone", "ipad", "android", "xiaomi", "mobile"]):
        return "Mobile Device", "Mobile Device Risk"
    if m:
        return "Network Device", "Network Device Risk"
    return "Unknown Asset", "General Asset Risk"


def build_asset_inventory(findings: pd.DataFrame, os_df: pd.DataFrame, adv_df: pd.DataFrame) -> pd.DataFrame:
    """Build a full active-asset inventory from Advanced IP Scanner, plus unmatched Nessus-only hosts."""
    host_os = build_host_os_map(os_df) if not os_df.empty else {}
    inventory_rows = []
    seen_ips: set[str] = set()
    known_names: set[str] = set()

    def add_asset(ip: str, name: str, manufacturer: str, mac: str, status: str, source: str) -> None:
        """Append one asset once, using IP as the primary identity."""
        ip = clean(ip, 100)
        name = clean(name, 300)
        manufacturer = clean(manufacturer, 300)
        mac = clean(mac, 100)
        status = clean(status, 100) or "Discovered"
        if not ip or ip in seen_ips:
            return

        os_name = host_os.get(ip, "") or host_os.get(name, "")
        classification, threat_category = classify_asset(name or ip, ip, os_name, manufacturer)
        inventory_rows.append({
            "Asset No.": len(inventory_rows) + 1,
            "Asset Name": name or ip,
            "IP Address": ip,
            "Department": "IT / Lab Environment",
            "Asset Classification": classification,
            "Threat Category": threat_category,
            "Manufacturer": manufacturer,
            "MAC Address": mac,
            "Status": status,
            "Registration Date": date.today().isoformat(),
            "Source": source,
        })
        seen_ips.add(ip)
        if name:
            known_names.add(name.casefold())
            # Only create a short-name alias for hostnames/FQDNs, never IP addresses.
            if not re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", name):
                known_names.add(name.split(".", 1)[0].casefold())

    # Primary inventory source: every active device found by Advanced IP Scanner.
    # This keeps devices with no Nessus vulnerabilities in the asset inventory.
    if not adv_df.empty:
        for _, row in adv_df.iterrows():
            add_asset(
                clean(row.get("IP", ""), 100),
                clean(row.get("Name", ""), 300),
                clean(row.get("Manufacturer", ""), 300),
                clean(row.get("MAC Address", ""), 100),
                clean(row.get("Status", ""), 100),
                "Advanced IP Scanner",
            )

    # Secondary source: add a Nessus host only when it is not already represented
    # in the Advanced IP Scanner inventory by IP, hostname, or short hostname.
    if not findings.empty and "Host" in findings.columns:
        unique_hosts = sorted({clean(x, 100) for x in findings["Host"].dropna().tolist() if clean(x, 100)})
        for host in unique_hosts:
            host_key = host.casefold()
            is_ip_host = bool(re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", host))
            short_host = host.split(".", 1)[0].casefold() if not is_ip_host else ""
            if host in seen_ips or host_key in known_names or (short_host and short_host in known_names):
                continue

            # Nessus-only hostname entries have no reliable IP in this report.
            # Keep them only when they truly are absent from the IP scan.
            if is_ip_host:
                add_asset(host, host, "", "", "Detected by Nessus", "Nessus Vulnerabilities report")
            else:
                # Preserve an unmatched hostname without manufacturing an IP address.
                os_name = host_os.get(host, "")
                classification, threat_category = classify_asset(host, host, os_name, "")
                inventory_rows.append({
                    "Asset No.": len(inventory_rows) + 1,
                    "Asset Name": host,
                    "IP Address": "",
                    "Department": "IT / Lab Environment",
                    "Asset Classification": classification,
                    "Threat Category": threat_category,
                    "Manufacturer": "",
                    "MAC Address": "",
                    "Status": "Detected by Nessus",
                    "Registration Date": date.today().isoformat(),
                    "Source": "Nessus Vulnerabilities report",
                })
                known_names.add(host_key)
                if short_host:
                    known_names.add(short_host)

    columns = [
        "Asset No.", "Asset Name", "IP Address", "Department",
        "Asset Classification", "Threat Category", "Manufacturer", "MAC Address", "Status",
        "Registration Date", "Source"
    ]
    return pd.DataFrame(inventory_rows, columns=columns)

def build_risk_matrix() -> pd.DataFrame:
    rows = []
    for likelihood_label, likelihood_score in RISK_LEVEL_SCORES.items():
        for impact_label, impact_score in RISK_LEVEL_SCORES.items():
            risk_score = likelihood_score * impact_score
            rows.append({
                "Likelihood": likelihood_label,
                "Likelihood Score": likelihood_score,
                "Impact": impact_label,
                "Impact Score": impact_score,
                "Risk Score": risk_score,
                "Risk Rating": risk_rating_from_score(risk_score),
            })
    return pd.DataFrame(rows)



# -------------------------
# Final output column cleanup
# -------------------------
def drop_existing_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Drop columns only if they exist, so the script stays safe if a sheet changes."""
    if df is None or df.empty:
        return df
    return df.drop(columns=[col for col in columns if col in df.columns])


def apply_requested_column_changes(
    asset_df: pd.DataFrame,
    threat_df: pd.DataFrame,
    treatment_df: pd.DataFrame,
    risk_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Apply the final column structure requested for the four main deliverables.

    Asset Inventory:
      - Remove IP/Hostname if present
      - Remove Operating System, based on the earlier decision to avoid unmatched OS data

    Threat Catalogue:
      - Remove Operating System and Department if present

    Risk Treatment Plan:
      - Remove Owner, Comments, Reassessment Date, and Treatment Measure

    Information Security Risk Assessment:
      - Remove Operating System, Vulnerability Name, Likelihood, Impact,
        Risk Calculation, Evidence / Source, Severity, and Comments
      - Rename Risk Score to Risk Score (Likelihood x Impact)
    """
    asset_df = drop_existing_columns(asset_df, [
        "IP/Hostname",
        "Operating System",
        "Department",
        "Status",
        "Source",
        "Manufacturer",
        "MAC Address",
    ])

    threat_df = drop_existing_columns(threat_df, [
        "Operating System",
        "Department",
    ])

    treatment_df = drop_existing_columns(treatment_df, [
        "Owner",
        "Comments",
        "Treatment Measure",
        "Risk Calculation",
        "Treatment Action",
        "Implementation Verification",
        "Reassessment Required",
        "Status",
    ])

    risk_df = drop_existing_columns(risk_df, [
        "Operating System",
        "Vulnerability Name",
        "Likelihood",
        "Impact",
        "Risk Calculation",
        "Evidence / Source",
        "Severity",
        "Comments",
    ])

    if risk_df is not None and not risk_df.empty and "Risk Score" in risk_df.columns:
        risk_df = risk_df.rename(columns={"Risk Score": "Risk Score (Likelihood x Impact)"})

    return asset_df, threat_df, treatment_df, risk_df

# -------------------------
# Excel writing and styling
# -------------------------
def style_workbook(path: Path) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [cell.value for cell in ws[1]]
        header_map = {header: index + 1 for index, header in enumerate(headers) if header}

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border

        for column_name in ["Severity", "Risk Rating"]:
            if column_name in header_map:
                column_index = header_map[column_name]
                for row_index in range(2, ws.max_row + 1):
                    cell = ws.cell(row_index, column_index)
                    value = clean(cell.value, 100)
                    key = "Info" if value == "Informational" else normalize_severity(value)
                    color = SEVERITY_COLORS.get(key)
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)
                        if key in ["Critical", "High"]:
                            cell.font = Font(color="FFFFFF", bold=True)

        for column_name in ["Plugin URL", "Plugin Link"]:
            if column_name in header_map:
                column_index = header_map[column_name]
                for row_index in range(2, ws.max_row + 1):
                    cell = ws.cell(row_index, column_index)
                    value = clean(cell.value, 500)
                    if value.startswith("http"):
                        cell.hyperlink = value
                        cell.font = Font(color="0000EE", underline="single")

        for index, column_cells in enumerate(ws.columns, start=1):
            max_len = max([len(clean(cell.value, 200)) for cell in list(column_cells)[:150]] + [10])
            ws.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 60)

    wb.save(path)


def write_output_files(
    prefix: str,
    findings: pd.DataFrame,
    risk_df: pd.DataFrame,
    treatment_df: pd.DataFrame,
    threat_df: pd.DataFrame,
    asset_df: pd.DataFrame,
    os_df: pd.DataFrame,
    adv_df: pd.DataFrame,
) -> tuple[Path, Path, Path, Path]:
    asset_path = Path(f"{prefix}_Asset_Inventory.xlsx")
    threat_path = Path(f"{prefix}_Threat_Catalogue.xlsx")
    treatment_path = Path(f"{prefix}_Risk_Treatment_Plan.xlsx")
    risk_path = Path(f"{prefix}_Information_Security_Risk_Assessment.xlsx")

    with pd.ExcelWriter(asset_path, engine="openpyxl") as writer:
        asset_df.to_excel(writer, sheet_name="Asset Inventory", index=False)
        if not os_df.empty:
            drop_existing_columns(os_df.copy(), ["Unsupported"]).to_excel(
                writer, sheet_name="Operating Systems", index=False
            )

    with pd.ExcelWriter(risk_path, engine="openpyxl") as writer:
        build_risk_criteria().to_excel(writer, sheet_name="Risk Criteria", index=False)
        build_risk_matrix().to_excel(writer, sheet_name="Risk Matrix", index=False)
        risk_df.to_excel(writer, sheet_name="Risk Assessment", index=False)

    with pd.ExcelWriter(treatment_path, engine="openpyxl") as writer:
        treatment_df.to_excel(writer, sheet_name="Risk Treatment Plan", index=False)

    with pd.ExcelWriter(threat_path, engine="openpyxl") as writer:
        threat_df.to_excel(writer, sheet_name="Threat Catalogue", index=False)

    style_workbook(asset_path)
    style_workbook(threat_path)
    style_workbook(treatment_path)
    style_workbook(risk_path)

    return asset_path, threat_path, treatment_path, risk_path


# -------------------------
# Main CLI
# -------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Generate four English GRC Excel deliverables from Nessus plain Vulnerabilities, OS, and Advanced IP Scanner reports.")
    parser.add_argument("--vulnerabilities", required=True, help="Nessus plain Vulnerabilities HTML report")
    parser.add_argument("--os", required=False, help="Nessus Operating Systems HTML report")
    parser.add_argument("--advip", required=False, help="Advanced IP Scanner HTML report")
    parser.add_argument("--risk-overrides", required=False, help="Optional JSON file with organization-specific per-plugin likelihood/impact overrides")
    parser.add_argument("--out-prefix", default="RAKIP", help="Output file prefix")
    args = parser.parse_args()

    vulnerabilities_path = Path(args.vulnerabilities)
    if not vulnerabilities_path.exists():
        raise SystemExit(f"[ERROR] File not found: {vulnerabilities_path}")

    print(f"[+] Parsing plain Nessus Vulnerabilities report: {vulnerabilities_path}")
    findings = parse_nessus_vulnerabilities(vulnerabilities_path)
    print(f"[+] Total Nessus findings extracted: {len(findings)}")

    os_df = parse_operating_systems(Path(args.os)) if args.os else pd.DataFrame()
    findings = enrich_findings_with_os(findings, os_df)

    adv_df = parse_advanced_ip_scanner(Path(args.advip)) if args.advip else pd.DataFrame()
    risk_overrides = load_risk_scoring_overrides(Path(args.risk_overrides)) if args.risk_overrides else {}

    scan_date = extract_scan_date(vulnerabilities_path)
    actionable_plugin_ids = findings[findings["Severity"].isin(["Critical", "High", "Medium", "Low"])]["Plugin ID"].astype(str).tolist()
    solution_map = fetch_tenable_solutions(actionable_plugin_ids)
    description_map = fetch_tenable_descriptions(actionable_plugin_ids)

    risk_df = build_risk_assessment(findings, description_map, risk_overrides)
    treatment_df = build_risk_treatment_plan(risk_df, findings, solution_map, scan_date)
    threat_df = build_threat_catalogue(findings)
    asset_df = build_asset_inventory(findings, os_df, adv_df)

    # Apply final requested column removals/renaming to the main deliverable sheets.
    asset_df, threat_df, treatment_df, risk_df = apply_requested_column_changes(
        asset_df, threat_df, treatment_df, risk_df
    )
    print(f"[+] Assets exported: {len(asset_df)}")
    print(f"[+] Threat catalogue entries exported: {len(threat_df)}")
    print(f"[+] Actionable risks exported: {len(risk_df)}")

    asset_path, threat_path, treatment_path, risk_path = write_output_files(args.out_prefix, findings, risk_df, treatment_df, threat_df, asset_df, os_df, adv_df)
    print(f"[OK] Created {asset_path}")
    print(f"[OK] Created {threat_path}")
    print(f"[OK] Created {treatment_path}")
    print(f"[OK] Created {risk_path}")


if __name__ == "__main__":
    main()
